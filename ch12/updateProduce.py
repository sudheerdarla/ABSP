#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {'Gralic': 3.07,
				 'Celery': 1.19,
				 'Lemon': 1.27}

# Todo: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): # skip the first row
	prodcueName = sheet.cell(row=rowNum, column=1).value
	if prodcueName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[prodcueName]

wb.save('updateProdcueSales.xlsx')